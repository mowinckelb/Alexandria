"""
Rebuild Numbers.xlsx.

Colour system (high intention density — every colour has a reason):
  Blue    — personal assumptions. The debatable numbers. The conversation.
  Green   — external facts. Verifiable. Stripe rates, existing costs, things that are what they are.
  Black   — calculated. Formula is the explanation. No colour needed.
  Grey    — notes/sources. Not the model, just context.

Bold is totals only. Italic is % lines (margin, growth).
Section headers: bold, grey fill, thin borders.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Colour system ---
BLUE = '0000CC'     # personal assumptions — the debate
GREEN = '339933'    # external facts — verifiable
GREY = '888888'     # notes

fn = Font(name='Calibri', size=11)                              # black — calculated
fi = Font(name='Calibri', size=11, italic=True)                  # black italic — % lines
fb = Font(name='Calibri', size=11, bold=True)                    # black bold — totals only
fh = Font(name='Calibri', size=14, bold=True)                    # title
fl = Font(name='Calibri', size=11, bold=True)                    # section label (bold but not a total — used for section headers only)

fa = Font(name='Calibri', size=11, color=BLUE)                   # blue — assumption
fai = Font(name='Calibri', size=11, color=BLUE, italic=True)     # blue italic
fe = Font(name='Calibri', size=11, color=GREEN)                  # green — external fact
fs = Font(name='Calibri', size=11, color=GREY)                   # grey — notes

# Fills & borders
fill_section = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
thin = Side(style='thin', color='AAAAAA')
border_tb = Border(top=thin, bottom=thin)
border_b = Border(bottom=thin)
ar = Alignment(horizontal='right')

# Number formats — dash for zero
D = '"$"#,##0_);\\("$"#,##0\\);"\\u2013"_);@_)'
N = '#,##0_);\\(#,##0\\);"\\u2013"_);@_)'
P = '#,##0.0%_);\\(#,##0.0%\\);"\\u2013"_);@_)'
P0 = '#,##0%_);\\(#,##0%\\);"\\u2013"_);@_)'
M = '0.0"x"'

A = 'Assumptions!'


def w(ws, row, col, val, font=fn, fmt='General', fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font
    c.number_format = fmt
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border


def section(ws, row, label, max_col=4):
    for col in range(2, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill_section
        c.border = border_tb
    w(ws, row, 2, label, fl, fill=fill_section, border=border_tb)


def total_border(ws, row, max_col=4):
    for col in range(2, max_col + 1):
        ws.cell(row=row, column=col).border = border_b


# ============================================================
# ASSUMPTIONS
# ============================================================
ws = wb.active
ws.title = 'Assumptions'
ws.column_dimensions['A'].width = 1.5
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 50

r = 1; w(ws, r, 2, 'Alexandria Library, Inc \u2014 Financial Model', fh, border=border_b)

# --- Pricing ---
r = 3; section(ws, r, 'Pricing')
# Pricing tiers are external facts (green) — these are the published prices
r = 4; w(ws, r, 2, 'Sovereignty \u2014 with kin'); w(ws, r, 3, 5, fe, D, align=ar); w(ws, r, 4, '$/month', fs)
r = 5; w(ws, r, 2, 'Sovereignty \u2014 without kin'); w(ws, r, 3, 10, fe, D, align=ar); w(ws, r, 4, '$/month', fs)
r = 6; w(ws, r, 2, 'Examined Life \u2014 with kin'); w(ws, r, 3, 15, fe, D, align=ar); w(ws, r, 4, '$/month', fs)
r = 7; w(ws, r, 2, 'Examined Life \u2014 without kin'); w(ws, r, 3, 20, fe, D, align=ar); w(ws, r, 4, '$/month', fs)
# Kin % is an assumption (blue) — we don't know this yet
r = 8; w(ws, r, 2, '% of users with 3+ active kin', fi); w(ws, r, 3, 0.6, fai, P, align=ar)
# Effective ARPU: blue — depends on discount assumptions
r = 9; w(ws, r, 2, 'Effective Sovereignty ARPU', fb); w(ws, r, 3, 6, fa, D, align=ar)
w(ws, r, 4, 'Blended $7, after quarterly/annual discounts', fs)
r = 10; w(ws, r, 2, 'Effective Examined Life ARPU', fb); w(ws, r, 3, 15, fa, D, align=ar)
w(ws, r, 4, 'Blended $17, after discounts', fs)
total_border(ws, 10)

# --- Payment Processing ---
r = 12; section(ws, r, 'Payment Processing')
# These are external facts (green) — published rates
r = 13; w(ws, r, 2, 'ACH rate (Sovereignty)'); w(ws, r, 3, 0.008, fe, P, align=ar); w(ws, r, 4, '0.8% flat, no per-txn fee', fs)
r = 14; w(ws, r, 2, 'Stripe % rate (Examined Life)'); w(ws, r, 3, 0.029, fe, P, align=ar); w(ws, r, 4, '2.9%', fs)
r = 15; w(ws, r, 2, 'Stripe per-txn fee'); w(ws, r, 3, 0.30, fe, D, align=ar); w(ws, r, 4, '$0.30/txn', fs)

# --- Growth ---
r = 17; section(ws, r, 'Growth')
# All growth assumptions are blue — the core debate
r = 18; w(ws, r, 2, 'Year 1 Sovereignty users (EOY)'); w(ws, r, 3, 2000, fa, N, align=ar)
r = 19; w(ws, r, 2, 'Year 2 growth'); w(ws, r, 3, 1.50, fai, P0, align=ar)
r = 20; w(ws, r, 2, 'Year 3 growth'); w(ws, r, 3, 1.20, fai, P0, align=ar)
r = 21; w(ws, r, 2, 'Year 4 growth'); w(ws, r, 3, 0.90, fai, P0, align=ar)
r = 22; w(ws, r, 2, 'Year 5 growth'); w(ws, r, 3, 0.60, fai, P0, align=ar)
r = 23; w(ws, r, 2, 'Year 1 Sov \u2192 EL conversion'); w(ws, r, 3, 0.08, fai, P, align=ar)
r = 24; w(ws, r, 2, 'Annual conversion improvement'); w(ws, r, 3, 0.02, fai, P, align=ar); w(ws, r, 4, '+2pp/year', fs)
r = 25; w(ws, r, 2, 'EL direct users (Year 1 seed)'); w(ws, r, 3, 50, fa, N, align=ar)
r = 26; w(ws, r, 2, 'EL organic growth'); w(ws, r, 3, 0.30, fai, P0, align=ar); w(ws, r, 4, '30% YoY', fs)
r = 27; w(ws, r, 2, 'Monthly churn \u2014 Sovereignty', fi); w(ws, r, 3, 0.04, fai, P, align=ar)
r = 28; w(ws, r, 2, 'Monthly churn \u2014 Examined Life', fi); w(ws, r, 3, 0.02, fai, P, align=ar)

# --- Company Opex ---
r = 30; section(ws, r, 'Company Opex (Monthly)')
# These are green — current known costs
r = 31; w(ws, r, 2, 'Claude Max'); w(ws, r, 3, 100, fe, D, align=ar); w(ws, r, 4, 'anthropic.com', fs)
r = 32; w(ws, r, 2, 'Fly.io'); w(ws, r, 3, 2, fe, D, align=ar); w(ws, r, 4, 'fly.io', fs)
r = 33; w(ws, r, 2, 'Total monthly opex', fb); w(ws, r, 3, '=C31+C32', fn, D, align=ar)
total_border(ws, 33)
r = 34; w(ws, r, 2, 'Annual opex', fb); w(ws, r, 3, '=C33*12', fn, D, align=ar)
r = 35; w(ws, r, 2, 'Free stack', fi); w(ws, r, 4, 'GitHub, Google Drive, Vercel, UptimeRobot, Google Cloud, Claude Code \u2014 all $0', fs)

# --- Legal ---
r = 37; section(ws, r, 'Legal & Filing Costs')
# Green — known costs, already incurred or published rates
r = 38; w(ws, r, 2, 'Stripe Atlas (Delaware C-Corp)'); w(ws, r, 3, 500, fe, D, align=ar); w(ws, r, 4, 'Paid', fs)
r = 39; w(ws, r, 2, 'California foreign qualification'); w(ws, r, 3, 150, fe, D, align=ar)
r = 40; w(ws, r, 2, 'Total one-time', fb); w(ws, r, 3, '=C38+C39', fn, D, align=ar)
total_border(ws, 40)
r = 41; w(ws, r, 2, 'Delaware franchise tax (annual)'); w(ws, r, 3, 400, fe, D, align=ar)
r = 42; w(ws, r, 2, 'California franchise tax (annual)'); w(ws, r, 3, 800, fe, D, align=ar)
r = 43; w(ws, r, 2, 'Registered agent (annual)'); w(ws, r, 3, 150, fe, D, align=ar)
# CPA costs are estimates — blue
r = 44; w(ws, r, 2, 'CPA \u2014 federal return'); w(ws, r, 3, 1500, fa, D, align=ar); w(ws, r, 4, 'Estimate', fs)
r = 45; w(ws, r, 2, 'CPA \u2014 state return(s)'); w(ws, r, 3, 500, fa, D, align=ar); w(ws, r, 4, 'Estimate', fs)
r = 46; w(ws, r, 2, 'Total recurring annual', fb); w(ws, r, 3, '=SUM(C41:C45)', fn, D, align=ar)
total_border(ws, 46)

# --- Founder ---
r = 48; section(ws, r, 'Founder Living Costs (not company opex)')
# Green — actual current costs
r = 49; w(ws, r, 2, 'Monthly'); w(ws, r, 3, 300, fe, D, align=ar); w(ws, r, 4, '$55/week food + T-Mobile + Apple One', fs)
r = 50; w(ws, r, 2, 'Annual', fb); w(ws, r, 3, '=C49*12', fn, D, align=ar)

# --- Marketing ---
r = 52; section(ws, r, 'Marketing')
# Blue — assumptions
r = 53; w(ws, r, 2, 'Year 1 ad spend (monthly)'); w(ws, r, 3, 500, fa, D, align=ar)
r = 54; w(ws, r, 2, 'Year 2+ ad spend (monthly)'); w(ws, r, 3, 1000, fa, D, align=ar)

# --- Library ---
r = 56; section(ws, r, 'Library for Labs (Year 3+)')
# Blue — assumptions about future B2B
r = 57; w(ws, r, 2, 'Average annual contract value'); w(ws, r, 3, 50000, fa, D, align=ar)
r = 58; w(ws, r, 2, 'Year 3 contracts'); w(ws, r, 3, 2, fa, N, align=ar)
r = 59; w(ws, r, 2, 'Year 4 contracts'); w(ws, r, 3, 5, fa, N, align=ar)
r = 60; w(ws, r, 2, 'Year 5 contracts'); w(ws, r, 3, 12, fa, N, align=ar)
r = 61; w(ws, r, 2, 'Library for People (Y4+, annual)'); w(ws, r, 3, 25000, fa, D, align=ar)

# --- Exit ---
r = 63; section(ws, r, 'Exit / Valuation')
# Blue — assumptions about market multiples
r = 64; w(ws, r, 2, 'Revenue multiple (SaaS)'); w(ws, r, 3, 10, fa, M, align=ar)
r = 65; w(ws, r, 2, 'Strategic premium (dataset)'); w(ws, r, 3, 2, fa, M, align=ar)
r = 66; w(ws, r, 2, 'Year 5 exit multiple', fb); w(ws, r, 3, '=C64*C65', fn, M, align=ar)
total_border(ws, 66)


# ============================================================
# 5-YEAR PROJECTIONS
# ============================================================
ws2 = wb.create_sheet('5-Year Projections')
ws2.column_dimensions['A'].width = 1.5
ws2.column_dimensions['B'].width = 40
ws2.column_dimensions['C'].width = 14
for col in ['D', 'E', 'F', 'G', 'H']:
    ws2.column_dimensions[col].width = 14

YCOLS = [4, 5, 6, 7, 8]  # D-H
MC = 8
years = ['Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5']
growth_cells = ['C19', 'C20', 'C21', 'C22']  # Y2-Y5 growth refs


def subsection(ws, row, label):
    for col in range(2, MC + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill_section
        c.border = border_b
    w(ws, row, 2, label, fl, fill=fill_section, border=border_b)


r = 1; w(ws2, r, 2, 'Alexandria Library, Inc \u2014 5-Year Projections', fh, border=border_b)

# Year headers
r = 3
for i, yr in enumerate(years):
    w(ws2, r, YCOLS[i], yr, fn, align=ar, border=border_tb)

# --- USERS ---
r = 4; subsection(ws2, r, 'Users')

r = 5; w(ws2, r, 2, 'Sovereignty users (EOY)')
w(ws2, r, YCOLS[0], f'={A}C18', fn, N, align=ar)
for i in range(4):
    prev = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i+1], f'=ROUND({prev}5*(1+{A}{growth_cells[i]}),0)', fn, N, align=ar)

r = 6; w(ws2, r, 3, 'Sov \u2192 EL conversion', fi)
w(ws2, r, YCOLS[0], f'={A}C23', fi, P, align=ar)
for i in range(1, 5):
    prev = get_column_letter(YCOLS[i-1])
    w(ws2, r, YCOLS[i], f'={prev}6+{A}C24', fi, P, align=ar)

r = 7; w(ws2, r, 2, 'Converted to Examined Life')
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'=ROUND({col}5*{col}6,0)', fn, N, align=ar)

r = 8; w(ws2, r, 2, 'Examined Life (direct)')
w(ws2, r, YCOLS[0], f'={A}C25', fn, N, align=ar)
for i in range(1, 5):
    prev = get_column_letter(YCOLS[i-1])
    w(ws2, r, YCOLS[i], f'=ROUND({prev}8*(1+{A}C26),0)', fn, N, align=ar)

r = 9; w(ws2, r, 2, 'Total Examined Life users', fb)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}7+{col}8', fn, N, align=ar)

r = 10; w(ws2, r, 2, 'Total paying users', fb)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}5+{col}9', fn, N, align=ar)
total_border(ws2, 10, MC)

# --- REVENUE ---
r = 12; subsection(ws2, r, 'Revenue')

r = 13; w(ws2, r, 2, 'Sovereignty subscriptions')
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}5*{A}C9*12', fn, D, align=ar)

r = 14; w(ws2, r, 3, '% growth', fi)
for i in range(1, 5):
    col = get_column_letter(YCOLS[i])
    prev = get_column_letter(YCOLS[i-1])
    w(ws2, r, YCOLS[i], f'={col}13/{prev}13-1', fi, P, align=ar)

r = 15; w(ws2, r, 2, 'Examined Life subscriptions')
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}9*{A}C10*12', fn, D, align=ar)

r = 16; w(ws2, r, 2, 'Total subscription revenue', fb)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}13+{col}15', fn, D, align=ar)
total_border(ws2, 16, MC)

r = 18; w(ws2, r, 2, 'Library for Labs (B2B)')
for i, ref in enumerate([None, None, 'C58', 'C59', 'C60']):
    if ref is None:
        w(ws2, r, YCOLS[i], 0, fn, D, align=ar)
    else:
        w(ws2, r, YCOLS[i], f'={A}{ref}*{A}C57', fn, D, align=ar)

r = 19; w(ws2, r, 2, 'Library for People')
for i in range(3):
    w(ws2, r, YCOLS[i], 0, fn, D, align=ar)
w(ws2, r, YCOLS[3], f'={A}C61', fn, D, align=ar)
w(ws2, r, YCOLS[4], f'={A}C61*2', fn, D, align=ar)

r = 20; w(ws2, r, 2, 'Total Library revenue', fb)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}18+{col}19', fn, D, align=ar)

r = 22; w(ws2, r, 2, 'TOTAL REVENUE', fb)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}16+{col}20', fn, D, align=ar)
total_border(ws2, 22, MC)

r = 23; w(ws2, r, 3, '% growth', fi)
for i in range(1, 5):
    col = get_column_letter(YCOLS[i])
    prev = get_column_letter(YCOLS[i-1])
    w(ws2, r, YCOLS[i], f'={col}22/{prev}22-1', fi, P, align=ar)

# --- COSTS ---
r = 25; subsection(ws2, r, 'Costs')

r = 26; w(ws2, r, 2, 'Infrastructure')
for i in range(5):
    w(ws2, r, YCOLS[i], f'={A}C34', fn, D, align=ar)

r = 27; w(ws2, r, 2, 'Payment processing')
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i],
      f'=ROUND({col}13*{A}C13+{col}15*{A}C14+{col}9*12*{A}C15,0)',
      fn, D, align=ar)

r = 28; w(ws2, r, 2, 'Taxes, permits, filings')
w(ws2, r, YCOLS[0], f'={A}C40+{A}C46', fn, D, align=ar)
for i in range(1, 5):
    w(ws2, r, YCOLS[i], f'={A}C46', fn, D, align=ar)

r = 29; w(ws2, r, 2, 'Ads / marketing')
w(ws2, r, YCOLS[0], f'={A}C53*12', fn, D, align=ar)
for i in range(1, 5):
    w(ws2, r, YCOLS[i], f'={A}C54*12', fn, D, align=ar)

r = 30; w(ws2, r, 2, 'Founder living costs', fi)
for i in range(5):
    w(ws2, r, YCOLS[i], f'={A}C50', fi, D, align=ar)

r = 31; w(ws2, r, 2, 'Total costs', fb)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'=SUM({col}26:{col}30)', fn, D, align=ar)
total_border(ws2, 31, MC)

# --- PROFITABILITY ---
r = 33; subsection(ws2, r, 'Profitability')

r = 34; w(ws2, r, 2, 'Net income', fb)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}22-{col}31', fn, D, align=ar)

r = 35; w(ws2, r, 3, '% margin', fi)
for i in range(5):
    col = get_column_letter(YCOLS[i])
    w(ws2, r, YCOLS[i], f'={col}34/{col}22', fi, P, align=ar)

r = 36; w(ws2, r, 2, 'Cumulative net income')
w(ws2, r, YCOLS[0], f'={get_column_letter(YCOLS[0])}34', fn, D, align=ar)
for i in range(1, 5):
    col = get_column_letter(YCOLS[i])
    prev = get_column_letter(YCOLS[i-1])
    w(ws2, r, YCOLS[i], f'={prev}36+{col}34', fn, D, align=ar)
total_border(ws2, 36, MC)

# --- VALUATION ---
r = 38; subsection(ws2, r, 'Valuation')

r = 39; w(ws2, r, 2, 'Year 5 ARR', fb)
w(ws2, r, YCOLS[4], f'={get_column_letter(YCOLS[4])}22', fn, D, align=ar)

r = 40; w(ws2, r, 2, 'Exit multiple')
w(ws2, r, YCOLS[4], f'={A}C66', fn, M, align=ar)

r = 41; w(ws2, r, 2, 'Implied valuation', fb)
c5 = get_column_letter(YCOLS[4])
w(ws2, r, YCOLS[4], f'={c5}39*{c5}40', fn, D, align=ar)
total_border(ws2, 41, MC)

# Save
wb.save('C:/Users/USER/Alexandria/files/confidential/Numbers.xlsx')
print('Rebuilt Numbers.xlsx')
print('Blue = assumptions (the debate). Green = external facts. Black = calculated. Grey = notes.')
